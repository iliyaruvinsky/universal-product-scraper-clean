#!/usr/bin/env python3
"""
Excel Output Validation Module
===============================
Standalone validation tool for analyzing scraped Excel outputs.
Compares original product names (Column B) with scraped names (Column E) in פירוט tab.
Uses the same scoring system from OPTION_1_DETAILED_FLOW.md.

Usage:
    python excel_validator.py path/to/excel_file.xlsx
    python excel_validator.py --threshold 7.0 path/to/excel_file.xlsx
    python excel_validator.py --batch output/*.xlsx

Author: Skywind Universal Scraper Team
Date: August 14, 2025
"""

import sys
import os
import re
import argparse
from pathlib import Path
from typing import Dict, List, Tuple, Optional
from datetime import datetime
import openpyxl

# Add src to path for imports
sys.path.append(os.path.join(os.getcwd(), 'src'))
from validation.scoring_engine import ProductScoringEngine
from dataclasses import dataclass

# Fix for Windows Unicode issues
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')


@dataclass
class ValidationResult:
    """Result of validating a single product row"""
    row_number: int
    original_name: str
    scraped_name: str
    scraped_clean: str  # After removing Hebrew
    score: float
    max_score: float
    status: str  # "VALID" or "REVIEW"
    issues: List[str]
    gates_passed: Dict[str, bool]


class ExcelValidator:
    """
    Validates Excel output using OPTION_1 scoring system.
    Compares original product names with scraped names.
    """
    
    def __init__(self, threshold: float = 8.0):
        """
        Initialize validator with score threshold.
        
        Args:
            threshold: Minimum score for valid match (default 8.0/10.0 = 80%)
        """
        self.threshold = threshold
        self.scoring_engine = ProductScoringEngine()
        self.validation_results = []
        self.summary_stats = {
            'total': 0,
            'valid': 0,
            'review': 0,
            'skipped': 0
        }
    
    def remove_hebrew_characters(self, text: str) -> str:
        """
        Remove Hebrew characters and clean the text.
        Keeps English letters, numbers, and common punctuation.
        Also handles known Hebrew product terms that should be translated.
        
        Args:
            text: Text potentially containing Hebrew characters
            
        Returns:
            Cleaned text with only non-Hebrew characters
        """
        if not text:
            return ""
        
        # Common Hebrew manufacturer names and their English equivalents
        hebrew_to_english = {
            'טורנדו': 'TORNADO',
            'אלקטרה': 'ELECTRA',
            'תדיראן': 'TADIRAN',
            'טורנאדו': 'TORNADO',  # Alternative spelling
            # Add more as needed
        }
        
        # Replace known Hebrew terms with English equivalents
        result = text
        for hebrew, english in hebrew_to_english.items():
            if hebrew in result:
                result = result.replace(hebrew, english)
        
        # Keep only ASCII letters, numbers, and common punctuation
        # This effectively removes remaining Hebrew (Unicode range 0x0590-0x05FF)
        cleaned = re.sub(r'[^\x00-\x7F]+', ' ', result)
        
        # Clean up extra spaces
        cleaned = ' '.join(cleaned.split())
        
        return cleaned.strip()
    
    def parse_product_components(self, product_name: str) -> Dict:
        """
        Parse product into components (EXACT COPY from working zap_scraper.py).
        OPTION_1_DETAILED_FLOW.md Phase 2.3
        
        Args:
            product_name: Product name to parse
            
        Returns:
            Dictionary with manufacturer, series, and model components
        """
        # Extract model number using documented regex
        model_match = re.search(r'(\d+(?:\/\d+[A-Z]*)?)', product_name)
        model_number = model_match.group(1) if model_match else ""
        
        # Extract manufacturer (first word)
        words = product_name.split()
        manufacturer = words[0] if words else ""
        
        # Extract series words (everything except manufacturer and model)
        series_words = []
        for word in words[1:]:  # Skip manufacturer
            if word != model_number and not re.match(r'^\d+(?:\/\d+[A-Z]*)?$', word):
                series_words.append(word.upper())
        
        return {
            'manufacturer': manufacturer.upper(),
            'series': series_words,
            'model': model_number
        }
    
    def parse_product_components_enhanced(self, product_name: str, target_model: str) -> Dict:
        """
        Enhanced parsing that prefers model numbers matching the target.
        Handles cases like '4 כ"ס WD Pro SQ Inv 45' where target is '45'.
        
        Args:
            product_name: Product name to parse
            target_model: Target model number from original product
            
        Returns:
            Dictionary with manufacturer, series, and model components
        """
        # Find all potential model numbers
        all_numbers = re.findall(r'(\d+(?:\/\d+[A-Z]*)?)', product_name)
        
        # Enhanced model selection: prefer target match
        model_number = ""
        if target_model in all_numbers:
            # Perfect match found - use it
            model_number = target_model
        elif all_numbers:
            # Fallback to original OPTION_1 logic (first number)
            model_number = all_numbers[0]
        
        # Extract manufacturer (first word)
        words = product_name.split()
        manufacturer = words[0] if words else ""
        
        # Extract series words (everything except manufacturer and model)
        series_words = []
        for word in words[1:]:  # Skip manufacturer
            if word != model_number and not re.match(r'^\d+(?:\/\d+[A-Z]*)?$', word):
                series_words.append(word.upper())
        
        return {
            'manufacturer': manufacturer.upper(),
            'series': series_words,
            'model': model_number
        }
    
    def calculate_validation_score(self, original_name: str, scraped_name: str) -> Tuple[float, Dict[str, bool], List[str]]:
        """
        Calculate validation score using OPTION_1 scoring system.
        Enhanced to handle multiple model numbers in scraped text.
        
        Args:
            original_name: Original product name from source
            scraped_name: Scraped product name (already cleaned of Hebrew)
            
        Returns:
            Tuple of (score, gates_passed, issues)
        """
        issues = []
        gates_passed = {
            'model_number': False,
            'product_type': False
        }
        
        # Parse original product (standard way)
        original_components = self.parse_product_components(original_name)
        
        # Enhanced parsing for scraped product (handle multiple numbers)
        scraped_components = self.parse_product_components_enhanced(scraped_name, original_components['model'])
        
        # PHASE 3: CRITICAL GATES (from OPTION_1_DETAILED_FLOW.md)
        
        # Gate 1: Model Number Gate (Phase 3.1.2)
        if original_components['model'] != scraped_components['model']:
            issues.append(f"Model mismatch: '{original_components['model']}' ≠ '{scraped_components['model']}'")
            return 0.0, gates_passed, issues  # DISQUALIFIED
        
        gates_passed['model_number'] = True
        
        # Gate 2: Product Type Gate (Phase 3.1.3)
        # Check for INV/INVERTER presence
        original_has_inv = any(word in ['INV', 'INVERTER'] for word in original_components['series'])
        scraped_text_upper = scraped_name.upper()
        scraped_has_inv = 'INV' in scraped_text_upper or 'INVERTER' in scraped_text_upper
        
        if original_has_inv and not scraped_has_inv:
            issues.append("Missing INV/INVERTER in scraped name")
            return 0.0, gates_passed, issues  # DISQUALIFIED
        
        gates_passed['product_type'] = True
        
        # PHASE 4: COMPONENT SCORING (from OPTION_1_DETAILED_FLOW.md Phase 4)
        total_score = 0.0
        
        # 4.1.1: Manufacturer Scoring (0-1.0 points = 10%)
        if original_components['manufacturer'] in scraped_text_upper:
            manufacturer_score = 1.0
        else:
            manufacturer_score = 0.0
            issues.append(f"Manufacturer '{original_components['manufacturer']}' not found")
        
        total_score += manufacturer_score
        
        # 4.1.2: Model Name Scoring (0-4.0 points = 40%)
        if original_components['series']:
            series_matches = 0
            missing_series = []
            
            for word in original_components['series']:
                # Handle INV ≡ INVERTER equivalence
                if word == 'INV' and 'INVERTER' in scraped_text_upper:
                    series_matches += 1
                elif word == 'INVERTER' and 'INV' in scraped_text_upper:
                    series_matches += 1
                elif word in scraped_text_upper:
                    series_matches += 1
                else:
                    missing_series.append(word)
            
            series_percentage = series_matches / len(original_components['series'])
            series_score = series_percentage * 4.0
            
            if missing_series:
                issues.append(f"Missing series components: {', '.join(missing_series)}")
        else:
            series_score = 4.0  # No series words to match
        
        total_score += series_score
        
        # 4.1.3: Model Number Scoring (0-5.0 points = 50%)
        model_score = 5.0  # Already passed gate, gets full points
        total_score += model_score
        
        # 4.1.4: Extra Words Penalty (minor)
        # We're lenient here since scraped names often have extra descriptors
        scraped_words = scraped_text_upper.split()
        original_words = [original_components['manufacturer']] + original_components['series'] + [original_components['model']]
        
        # Count truly unrelated words (not partial matches)
        extra_words = []
        for word in scraped_words:
            if not any(orig in word or word in orig for orig in original_words):
                # Skip common ZAP additions
                if word not in ['ZAP', 'SHOP', 'LOGO', 'מזגן', 'עילי']:
                    extra_words.append(word)
        
        extra_penalty = min(len(extra_words) * 0.1, 1.0)  # Cap at 1.0
        
        if extra_penalty > 0:
            issues.append(f"Extra words found: {', '.join(extra_words[:5])}")  # Show first 5
        
        final_score = max(0.0, total_score - extra_penalty)
        
        return final_score, gates_passed, issues
    
    def validate_excel_file(self, excel_path: str) -> bool:
        """
        Validate an Excel file by comparing original vs scraped product names.
        
        Args:
            excel_path: Path to Excel file to validate
            
        Returns:
            True if validation completed successfully
        """
        try:
            print(f"\n📊 Validating Excel file: {excel_path}")
            print("=" * 60)
            
            # Load the Excel file
            wb = openpyxl.load_workbook(excel_path, read_only=True)
            
            # Check if פירוט sheet exists
            if 'פירוט' not in wb.sheetnames:
                print("❌ Error: 'פירוט' sheet not found in Excel file")
                return False
            
            sheet = wb['פירוט']
            
            # Process each row (skip header row 1)
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                # Column B (index 1): Original product name
                # Column E (index 4): Scraped product name
                
                if row and len(row) > 4:
                    original_name = row[1] if row[1] else ""
                    scraped_name = row[4] if row[4] else ""
                    
                    if not original_name or not scraped_name:
                        self.summary_stats['skipped'] += 1
                        continue
                    
                    # Clean scraped name (remove Hebrew)
                    scraped_clean = self.remove_hebrew_characters(scraped_name)
                    
                    # Calculate validation score using ProductScoringEngine
                    scoring_result = self.scoring_engine.calculate_match_score(original_name, scraped_clean)
                    score = scoring_result.total_score
                    gates_passed = scoring_result.gates_passed  
                    issues = scoring_result.issues
                    
                    # Determine status
                    status = "VALID" if score >= self.threshold else "REVIEW"
                    
                    # Create validation result
                    result = ValidationResult(
                        row_number=row_idx,
                        original_name=original_name,
                        scraped_name=scraped_name,
                        scraped_clean=scraped_clean,
                        score=score,
                        max_score=10.0,
                        status=status,
                        issues=issues,
                        gates_passed=gates_passed
                    )
                    
                    self.validation_results.append(result)
                    self.summary_stats['total'] += 1
                    
                    if status == "VALID":
                        self.summary_stats['valid'] += 1
                    else:
                        self.summary_stats['review'] += 1
            
            wb.close()
            return True
            
        except Exception as e:
            print(f"❌ Error validating Excel file: {e}")
            return False
    
    def create_validation_worksheet(self, excel_path: str) -> bool:
        """
        Create a 3rd worksheet in the Excel file with validation results.
        Adds a 'אימות נתונים' (Data Validation) worksheet with rejected products.
        
        Args:
            excel_path: Path to Excel file to modify
            
        Returns:
            True if worksheet created successfully
        """
        try:
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
            
            # Open the Excel file in read-write mode
            wb = openpyxl.load_workbook(excel_path)
            
            # Check if validation worksheet already exists
            validation_sheet_name = 'אימות נתונים'
            if validation_sheet_name in wb.sheetnames:
                # Remove existing validation sheet
                wb.remove(wb[validation_sheet_name])
            
            # Create new validation worksheet
            ws = wb.create_sheet(validation_sheet_name)
            
            # Define headers (same as פירוט + validation columns)
            headers = [
                'מספר שורה מקור',      # A: Source Row
                'שם מוצר מקורי',       # B: Original Product Name  
                'מחיר רשמי',           # C: Official Price
                'שם ספק',             # D: Vendor Name
                'שם מוצר באתר ספק',     # E: Scraped Product Name
                'מחיר ספק',           # F: Vendor Price
                'הפרש מחיר',          # G: Price Difference
                'אחוז הפרש',           # H: Percentage Difference
                'קישור לספק',         # I: Vendor URL
                'חותמת זמן',          # J: Timestamp
                'ציון אימות',         # K: Validation Score
                'סטטוס',              # L: Status
                'סיבות דחייה'         # M: Rejection Reasons
            ]
            
            # Set headers
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red background
                cell.alignment = Alignment(horizontal="center")
            
            # Get rejected products
            rejected_results = [r for r in self.validation_results if r.status == "REVIEW"]
            
            if not rejected_results:
                # Add a note that no rejections were found
                ws.cell(row=2, column=1, value="✅ לא נמצאו מוצרים הדורשים בדיקה נוספת")
                ws.cell(row=2, column=1).font = Font(color="008000", bold=True)  # Green
            else:
                # Read original Excel data to get full context
                details_sheet = wb['פירוט']
                
                # Add rejected products
                row_num = 2
                for result in rejected_results:
                    # Get original row data from פירוט sheet
                    original_row = list(details_sheet[result.row_number])
                    
                    # Copy original data
                    for col in range(1, 11):  # Columns A-J from original
                        if col <= len(original_row):
                            ws.cell(row=row_num, column=col, value=original_row[col-1].value)
                    
                    # Add validation data
                    ws.cell(row=row_num, column=11, value=f"{result.score:.1f}/10.0")  # Validation Score
                    ws.cell(row=row_num, column=12, value="⚠️ דורש בדיקה")  # Status
                    
                    # Format rejection reasons nicely
                    reasons_text = " | ".join(result.issues) if result.issues else "ציון נמוך מהסף הנדרש"
                    ws.cell(row=row_num, column=13, value=reasons_text)  # Rejection Reasons
                    
                    # Color the row
                    for col in range(1, 14):
                        cell = ws.cell(row=row_num, column=col)
                        cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # Light yellow
                    
                    row_num += 1
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                # Set column width (with limits)
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save the updated Excel file
            wb.save(excel_path)
            wb.close()
            
            print(f"📊 Validation worksheet '{validation_sheet_name}' created with {len(rejected_results)} rejected products")
            return True
            
        except Exception as e:
            print(f"❌ Error creating validation worksheet: {e}")
            return False
    
    def generate_report(self, output_path: Optional[str] = None) -> str:
        """
        Generate validation report.
        
        Args:
            output_path: Optional path to save report to file
            
        Returns:
            Report text
        """
        report_lines = []
        
        # Header
        report_lines.append("=" * 70)
        report_lines.append("EXCEL VALIDATION REPORT")
        report_lines.append("=" * 70)
        report_lines.append(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        report_lines.append(f"Validation Threshold: {self.threshold:.1f}/10.0 ({self.threshold*10:.0f}%)")
        report_lines.append("")
        
        # Summary
        report_lines.append("SUMMARY")
        report_lines.append("-" * 30)
        report_lines.append(f"Total Products Validated: {self.summary_stats['total']}")
        
        if self.summary_stats['total'] > 0:
            valid_pct = (self.summary_stats['valid'] / self.summary_stats['total']) * 100
            review_pct = (self.summary_stats['review'] / self.summary_stats['total']) * 100
            
            report_lines.append(f"✅ Valid: {self.summary_stats['valid']} products ({valid_pct:.1f}%)")
            report_lines.append(f"⚠️  Review Needed: {self.summary_stats['review']} products ({review_pct:.1f}%)")
            
            if self.summary_stats['skipped'] > 0:
                report_lines.append(f"⏭️  Skipped (empty): {self.summary_stats['skipped']} rows")
        
        report_lines.append("")
        
        # Detailed findings for products needing review
        review_results = [r for r in self.validation_results if r.status == "REVIEW"]
        
        if review_results:
            report_lines.append("PRODUCTS REQUIRING REVIEW")
            report_lines.append("-" * 30)
            
            for result in review_results[:20]:  # Show first 20
                report_lines.append("")
                report_lines.append(f"⚠️  Row {result.row_number} - REVIEW NEEDED (Score: {result.score:.1f}/10.0)")
                report_lines.append(f"   Original: \"{result.original_name}\"")
                report_lines.append(f"   Scraped:  \"{result.scraped_name}\"")
                report_lines.append(f"   Cleaned:  \"{result.scraped_clean}\"")
                
                if result.issues:
                    report_lines.append(f"   Issues:")
                    for issue in result.issues:
                        report_lines.append(f"     • {issue}")
                
                # Show gate status
                gates_status = []
                if not result.gates_passed.get('model_number'):
                    gates_status.append("Model Gate FAILED")
                if not result.gates_passed.get('product_type'):
                    gates_status.append("Type Gate FAILED")
                
                if gates_status:
                    report_lines.append(f"   Gates: {', '.join(gates_status)}")
            
            if len(review_results) > 20:
                report_lines.append("")
                report_lines.append(f"... and {len(review_results) - 20} more products requiring review")
        
        # Sample of valid products (for verification)
        valid_results = [r for r in self.validation_results if r.status == "VALID"]
        
        if valid_results and len(valid_results) <= 10:
            report_lines.append("")
            report_lines.append("VALID PRODUCTS (Sample)")
            report_lines.append("-" * 30)
            
            for result in valid_results[:5]:
                report_lines.append("")
                report_lines.append(f"✅ Row {result.row_number} - VALID (Score: {result.score:.1f}/10.0)")
                report_lines.append(f"   Original: \"{result.original_name}\"")
                report_lines.append(f"   Scraped:  \"{result.scraped_clean}\"")
        
        # Footer
        report_lines.append("")
        report_lines.append("=" * 70)
        report_lines.append("END OF VALIDATION REPORT")
        report_lines.append("=" * 70)
        
        report_text = "\n".join(report_lines)
        
        # Save to file if requested
        if output_path:
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(report_text)
            print(f"\n📄 Report saved to: {output_path}")
        
        return report_text


def main():
    """Main entry point for command-line usage."""
    parser = argparse.ArgumentParser(
        description="Validate Excel output from Universal Product Scraper"
    )
    parser.add_argument(
        'excel_file',
        help='Path to Excel file to validate'
    )
    parser.add_argument(
        '--threshold',
        type=float,
        default=8.0,
        help='Minimum score threshold for valid match (default: 8.0)'
    )
    parser.add_argument(
        '--output',
        help='Path to save validation report'
    )
    parser.add_argument(
        '--verbose',
        action='store_true',
        help='Show detailed output during validation'
    )
    parser.add_argument(
        '--create-worksheet',
        action='store_true',
        help='Create validation worksheet in the Excel file with rejected products'
    )
    
    args = parser.parse_args()
    
    # Check if file exists
    if not Path(args.excel_file).exists():
        print(f"❌ Error: File not found: {args.excel_file}")
        sys.exit(1)
    
    # Create validator
    validator = ExcelValidator(threshold=args.threshold)
    
    # Validate the Excel file
    if validator.validate_excel_file(args.excel_file):
        # Generate and print report
        report = validator.generate_report(args.output)
        print(report)
        
        # Create validation worksheet if requested
        if args.create_worksheet:
            print("\n" + "="*50)
            print("📊 CREATING VALIDATION WORKSHEET")
            print("="*50)
            validator.create_validation_worksheet(args.excel_file)
        
        # Exit with appropriate code
        if validator.summary_stats['review'] > 0:
            sys.exit(1)  # Exit with error if products need review
        else:
            sys.exit(0)  # Success
    else:
        sys.exit(2)  # Validation error


if __name__ == "__main__":
    main()
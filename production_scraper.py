#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
PRODUCTION SCRAPER - FIXED UNIFIED VERSION
Unified extraction logic for both explicit and headless modes
Based on proven HTML structure comparison results
"""

import sys
import os
import time
import re
import argparse
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Set UTF-8 encoding
if sys.platform == "win32":
    sys.stdout.reconfigure(encoding='utf-8')

# Add src to path for imports
sys.path.append(os.path.join(os.getcwd(), 'src'))
from excel.source_reader import SourceExcelReader
from validation.scoring_engine import ProductScoringEngine

# Global headless mode flag
HEADLESS_MODE = False

def create_driver():
    """Create Chrome driver with unified configuration."""
    chrome_options = Options()
    chrome_options.add_argument('--no-sandbox')
    chrome_options.add_argument('--disable-dev-shm-usage')
    chrome_options.add_argument('--start-maximized')
    chrome_options.add_argument('--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36')
    
    if HEADLESS_MODE:
        chrome_options.add_argument('--headless')
        chrome_options.add_argument('--disable-background-timer-throttling')
        chrome_options.add_argument('--disable-backgrounding-occluded-windows')
        chrome_options.add_argument('--disable-renderer-backgrounding')
    
    driver = webdriver.Chrome(options=chrome_options)
    if HEADLESS_MODE:
        driver.set_window_size(1920, 1080)
    
    return driver

def search_and_filter_product(driver, product_name):
    """Search for product and navigate to model page using breakthrough method."""
    print(f"\n🔍 SEARCHING FOR: {product_name}")
    print(f"🤖 Mode: {'HEADLESS' if HEADLESS_MODE else 'VISIBLE'}")
    
    try:
        # Step 1: Navigate to ZAP homepage
        driver.get("https://zap.co.il")
        time.sleep(2)
        
        # Step 2: Find and use search box
        search_box = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "#acSearch-input"))
        )
        search_box.click()
        time.sleep(1)
        
        # Try hyphenated approach first (breakthrough method)
        hyphenated = product_name.replace(" ", "-")
        search_box.clear()
        search_box.send_keys(hyphenated)
        time.sleep(3)
        
        # Step 3: Look for dropdown suggestions
        dropdown_containers = driver.find_elements(By.CSS_SELECTOR, ".acSearch-row-container")
        
        if dropdown_containers:
            print(f"✅ Found {len(dropdown_containers)} dropdown suggestions")
            # Look for suggestions with model ID in data-search-link
            from src.hebrew.text_processor import HebrewTextProcessor
            hebrew_processor = HebrewTextProcessor()
            
            for container in dropdown_containers:
                try:
                    # First check if this is a valid HVAC suggestion
                    suggestion_text = container.text.strip()
                    
                    # Skip phone/mobile products
                    if hebrew_processor.contains_phone_keywords(suggestion_text):
                        print(f"🚫 Skipping phone suggestion: {suggestion_text[:50]}...")
                        continue
                    
                    # Only process HVAC products
                    if not hebrew_processor.contains_hvac_keywords(suggestion_text):
                        print(f"🚫 Skipping non-HVAC suggestion: {suggestion_text[:50]}...")
                        continue
                    
                    # Find the clickable div with data-search-link attribute
                    suggestion_div = container.find_element(By.CSS_SELECTOR, ".acSearch-row.acSearch-row-img[data-search-link*='modelid=']")
                    search_link = suggestion_div.get_attribute('data-search-link')
                    
                    # Extract model ID from the data-search-link
                    if 'modelid=' in search_link:
                        import re
                        model_id_match = re.search(r'modelid=(\d+)', search_link)
                        if model_id_match:
                            model_id = model_id_match.group(1)
                            # Construct direct model page URL
                            model_url = f"https://www.zap.co.il/model.aspx?modelid={model_id}"
                            print(f"✅ Found valid HVAC model ID {model_id} in dropdown")
                            print(f"📍 Navigating directly to: {model_url}")
                            
                            # Navigate directly to the model page
                            driver.get(model_url)
                            time.sleep(5 if HEADLESS_MODE else 3)
                            return "success", model_url
                except:
                    continue
            
            # Fallback: find valid HVAC suggestion if no model ID found
            try:
                from src.hebrew.text_processor import HebrewTextProcessor
                hebrew_processor = HebrewTextProcessor()
                
                for container in dropdown_containers:
                    try:
                        suggestion_text = container.text.strip()
                        
                        # Check if this is a phone/mobile product
                        if hebrew_processor.contains_phone_keywords(suggestion_text):
                            print(f"🚫 REJECTED phone dropdown: {suggestion_text[:50]}...")
                            continue
                        
                        # Check if this is an HVAC product
                        if hebrew_processor.contains_hvac_keywords(suggestion_text):
                            suggestion_div = container.find_element(By.CSS_SELECTOR, ".acSearch-row.acSearch-row-img")
                            print(f"✅ Clicking valid HVAC suggestion: {suggestion_text[:50]}...")
                            suggestion_div.click()
                            time.sleep(5 if HEADLESS_MODE else 3)
                            return "success", driver.current_url
                    except:
                        continue
                
                print(f"❌ No valid HVAC suggestions found in dropdown")
            except:
                pass
        
        # SUB-OPTION 1B: Fallback to space format if 1A failed
        print(f"🔍 SUB-OPTION 1A failed, trying SUB-OPTION 1B (space format)...")
        
        # Clear search box and enter space-separated format
        search_box = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "#acSearch-input"))
        )
        search_box.click()
        time.sleep(1)
        search_box.clear()
        search_box.send_keys(product_name)  # Use original space-separated format
        print(f"🔍 SUB-OPTION 1B: Entered space format: {product_name}")
        search_box.send_keys(Keys.ENTER)
        time.sleep(5 if HEADLESS_MODE else 3)
        
        # Step 5: Look for product links in search results - ENHANCED 2025 SELECTORS
        try:
            # ENHANCED: Use correct selectors based on HTML analysis 
            # Priority order: NEW selectors first (fs.aspx), legacy last
            product_link_selectors = [
                "a[href*='/fs.aspx']",             # NEW: Main product links from HTML analysis
                "a[href*='fs.aspx']",              # NEW: Alternative fs links
                ".ModelTitle",                     # NEW: Product title links
                ".ModelPic",                       # NEW: Product image links
                ".noModelRow.ModelRow a",          # NEW: Links in product containers
                "a[href*='model.aspx?modelid=']",  # Legacy selector (last resort)
            ]
            
            product_links = []
            for selector in product_link_selectors:
                try:
                    links = driver.find_elements(By.CSS_SELECTOR, selector)
                    if links:
                        print(f"✅ Found {len(links)} links with selector: {selector}")
                        # Filter out duplicate links and ads
                        filtered_links = []
                        for link in links:
                            href = link.get_attribute('href') or ''
                            text = link.text.strip()
                            
                            # Skip ad links
                            if 'data-bid-id' in link.get_attribute('outerHTML') or 'מודעה' in text:
                                continue
                            
                            # Skip duplicates
                            if href not in [l.get_attribute('href') for l in filtered_links]:
                                filtered_links.append(link)
                        
                        if filtered_links:
                            print(f"✅ Using {len(filtered_links)} filtered links from: {selector}")
                            product_links = filtered_links
                            break  # Use first successful selector with valid links
                except Exception as e:
                    print(f"❌ Selector {selector} failed: {e}")
                    continue
            
            if product_links:
                print(f"✅ Found {len(product_links)} total product links in results")
                
                # ENHANCED VALIDATION: Use scoring engine to find best match
                from src.hebrew.text_processor import HebrewTextProcessor
                from src.validation.scoring_engine import ProductScoringEngine
                
                hebrew_processor = HebrewTextProcessor()
                scoring_engine = ProductScoringEngine()
                
                # Track best match with scoring
                best_match = None
                best_score = 0.0
                best_match_text = ""
                
                for link in product_links:
                    try:
                        # ENHANCED: Extract text from different link types
                        link_text = link.text.strip()
                        
                        # NEW: For fs.aspx links, get text from nearby elements or data attributes
                        if '/fs.aspx' in link.get_attribute('href'):
                            # Try to get product name from aria-label or data attributes
                            aria_label = link.get_attribute('aria-label')
                            if aria_label:
                                link_text = aria_label.strip()
                            
                            # Try to get product text from ModelRow container
                            try:
                                model_row = link.find_element(By.XPATH, "./ancestor::*[contains(@class, 'ModelRow')]")
                                if model_row:
                                    model_text = model_row.text.strip()
                                    # Extract product name from model text (before price)
                                    product_match = model_text.split('₪')[0].strip()
                                    if len(product_match) > len(link_text):
                                        link_text = product_match
                            except:
                                pass
                        
                        parent_text = ""
                        try:
                            parent_text = link.find_element(By.XPATH, "..").text.strip()
                        except:
                            pass
                        
                        combined_text = f"{link_text} {parent_text}"
                        
                        # Check if this is a phone/mobile product
                        if hebrew_processor.contains_phone_keywords(combined_text):
                            print(f"🚫 REJECTED phone product: {link_text[:50]}...")
                            continue
                        
                        # Check if this is an HVAC product or contains Tornado
                        is_hvac = (hebrew_processor.contains_hvac_keywords(combined_text) or
                                  'tornado' in combined_text.lower() or
                                  'TORNADO' in combined_text or
                                  'טורנדו' in combined_text)
                        
                        if not is_hvac:
                            print(f"🚫 REJECTED non-HVAC product: {link_text[:50]}...")
                            continue
                        
                        # NEW: Score-based validation for HVAC products
                        scoring_result = scoring_engine.calculate_match_score(
                            product_name,  # Original search term
                            link_text      # Found product name
                        )
                        
                        print(f"📊 Evaluating HVAC match: {link_text[:50]}... Score: {scoring_result.total_score:.1f}/10")
                        
                        # Track best match
                        if scoring_result.total_score > best_score:
                            best_score = scoring_result.total_score
                            best_match = link
                            best_match_text = link_text
                            
                    except Exception as e:
                        print(f"⚠️ Error evaluating link: {e}")
                        continue
                
                # Accept best match only if above threshold (8.0/10 as per CLAUDE.md)
                if best_match and best_score >= 8.0:
                    print(f"✅ ACCEPTED best match (score {best_score:.1f}/10): {best_match_text[:50]}...")
                    print(f"📍 STAYING ON LISTINGS PAGE to extract vendors from current page")
                    # DON'T navigate away - extract vendors from current listings page
                    return "success", driver.current_url
                elif best_match:
                    print(f"❌ Best match scored {best_score:.1f}/10 (below 8.0 threshold): {best_match_text[:50]}...")
                    print(f"❌ No products met the validation threshold")
                    return "no_valid_products", driver.current_url
                else:
                    print(f"❌ No valid HVAC products found among {len(product_links)} results")
                    return "no_valid_products", driver.current_url
            else:
                print(f"❌ No product links found")
                return "no_products", driver.current_url
        except Exception as e:
            print(f"❌ Error finding product links: {e}")
            return "failed", driver.current_url
            
    except Exception as e:
        print(f"❌ Search failed: {e}")
        return "failed", driver.current_url

def extract_zap_product_name(driver):
    """Extract product name from ZAP page using working ZapScraper pattern."""
    try:
        # Try multiple selectors (copied from working ZapScraper._extract_product_name)
        selectors = [
            "h1", "h2.product-title",
            "[class*='product-name']", "[class*='product-title']",
            "[itemprop='name']", ".title", ".product_title"
        ]
        
        for selector in selectors:
            elements = driver.find_elements(By.CSS_SELECTOR, selector)
            for elem in elements:
                text = elem.text.strip()
                if text and 5 < len(text) < 200:
                    print(f"✅ Extracted ZAP product name: {text}")
                    return text
                    
    except Exception as e:
        print(f"⚠️ Could not extract ZAP product name: {e}")
        
    return "Unknown Product"

def extract_individual_vendor_product_name(driver, vendor_url, vendor_name):
    """Extract product name from individual vendor page (like working ZapScraper)."""
    if not vendor_url:
        return f"Unknown Product ({vendor_name})"
    
    # Save current window
    original_window = driver.current_window_handle
    
    try:
        # Open vendor page in new tab
        driver.execute_script("window.open('');")
        driver.switch_to.window(driver.window_handles[-1])
        
        print(f"   🔍 Visiting {vendor_name} page for product name...")
        
        # Set page load timeout to prevent hanging
        driver.set_page_load_timeout(10)
        try:
            driver.get(vendor_url)
        except:
            # If page load times out, try to work with what we have
            pass
        
        time.sleep(1)  # Reduced from 2 seconds
        
        # Extract product name using COMPREHENSIVE selectors
        selectors = [
            "h1", "h2", "h3",  # Any heading
            ".product-title", ".product-name", ".product_title", ".product_name",
            "[class*='product-title']", "[class*='product-name']", 
            "[class*='title']", "[class*='name']",
            "[itemprop='name']", "[data-product-name]",
            ".title", ".name", ".item-title", ".item-name",
            "title", "meta[property='og:title']"  # Page title and meta
        ]
        
        # Try selectors first
        for selector in selectors:
            try:
                elements = driver.find_elements(By.CSS_SELECTOR, selector)
                for elem in elements:
                    text = elem.text.strip() if hasattr(elem, 'text') else elem.get_attribute('content')
                    if text and 10 < len(text) < 300:  # Slightly relaxed length requirements
                        # Filter out obvious non-product text
                        text_lower = text.lower()
                        if not any(skip in text_lower for skip in ['cookie', 'policy', 'menu', 'nav', 'footer', 'header']):
                            print(f"   ✅ Extracted from {vendor_name}: {text}")
                            return text
            except:
                continue
        
        # Fallback: Try page title
        try:
            title = driver.title.strip()
            if title and 10 < len(title) < 300:
                print(f"   📄 Extracted from page title: {title}")
                return title
        except:
            pass
        
        print(f"   ⚠️ Could not extract product name from {vendor_name}")
        return f"Unknown Product ({vendor_name})"
        
    except Exception as e:
        print(f"   ❌ Failed to visit {vendor_name}: {e}")
        return f"Unknown Product ({vendor_name})"
        
    finally:
        # Close vendor tab and return to original
        try:
            driver.close()
            driver.switch_to.window(original_window)
        except:
            pass

def extract_vendors_unified(driver, zap_product_name="Unknown Product"):
    """UNIFIED vendor extraction - identical logic for both modes."""
    print(f"\n🏪 UNIFIED VENDOR EXTRACTION")
    print(f"🤖 Mode: {'HEADLESS' if HEADLESS_MODE else 'VISIBLE'}")
    
    vendors = []
    
    try:
        # ENHANCED 2025: Use correct selectors based on page type
        current_url = driver.current_url
        
        if "models.aspx" in current_url:
            # On listings page - use ModelRow selectors (same as product detection)
            vendor_rows = driver.find_elements(By.CSS_SELECTOR, ".noModelRow.ModelRow")
            print(f"✅ Found {len(vendor_rows)} vendor listings on models.aspx page")
        else:
            # On comparison page - use original selectors
            vendor_rows = driver.find_elements(By.CSS_SELECTOR, ".compare-item-row.product-item")
            print(f"✅ Found {len(vendor_rows)} vendor rows using comparison page selectors")
        
        if len(vendor_rows) == 0:
            # Enhanced fallback for both page types
            fallback_selectors = [
                ".compare-item-row",           # Original fallback
                ".ModelRow",                   # Listings page fallback
                ".noModelRow",                 # Alternative listings fallback
            ]
            
            for selector in fallback_selectors:
                vendor_rows = driver.find_elements(By.CSS_SELECTOR, selector)
                if vendor_rows:
                    print(f"✅ Fallback found {len(vendor_rows)} rows with: {selector}")
                    break
        
        for i, row in enumerate(vendor_rows, 1):
            try:
                print(f"\n🔍 Processing vendor {i}:")
                
                # ENHANCED: Extract vendor name based on page type
                vendor_name = f"Vendor {i}"
                
                if "models.aspx" in current_url:
                    # Extract vendor from listings page (models.aspx)
                    try:
                        # Look for vendor name in store name section
                        store_link = row.find_element(By.CSS_SELECTOR, "a[data-site-name]")
                        vendor_name = store_link.get_attribute('data-site-name')
                        print(f"   📍 Found vendor via data-site-name: {vendor_name}")
                    except:
                        try:
                            # Alternative: look for vendor in text content
                            store_text = row.find_element(By.CSS_SELECTOR, ".store-name, .short-store-name")
                            vendor_text = store_text.text.strip()
                            if 'ב-' in vendor_text:
                                vendor_name = vendor_text.split('ב-')[1].strip()
                                print(f"   📍 Found vendor via store text: {vendor_name}")
                        except:
                            try:
                                # Fallback: extract from any link title
                                store_link = row.find_element(By.CSS_SELECTOR, "a[title]")
                                vendor_name = store_link.get_attribute('title')
                                print(f"   📍 Found vendor via link title: {vendor_name}")
                            except:
                                print(f"   ⚠️ Could not extract vendor name from ModelRow")
                else:
                    # Original logic for comparison pages
                    try:
                        logo_img = row.find_element(By.CSS_SELECTOR, ".compare-item-image.store a img[title]")
                        vendor_title = logo_img.get_attribute('title')
                        if vendor_title and vendor_title.lower() not in ['bullet', 'icon']:
                            vendor_name = vendor_title
                    except:
                        try:
                            all_imgs = row.find_elements(By.CSS_SELECTOR, "img[title]")
                            for img in all_imgs:
                                title = img.get_attribute('title')
                                if title and title.lower() not in ['bullet', 'icon', ''] and len(title) > 2:
                                    vendor_name = title
                                    break
                        except:
                            pass
                
                # ENHANCED: Extract vendor URL based on page type
                vendor_url = ""
                button_text = "לפרטים נוספים"
                
                if "models.aspx" in current_url:
                    # Extract URL from listings page (models.aspx)
                    try:
                        # Look for fs.aspx links in ModelRow
                        vendor_link = row.find_element(By.CSS_SELECTOR, "a[href*='/fs.aspx']")
                        vendor_url = vendor_link.get_attribute('href')
                        button_text = vendor_link.text.strip() or "לפרטים נוספים"
                        print(f"   🔗 Found vendor URL: {vendor_url[:100]}...")
                    except:
                        try:
                            # Alternative: any fs link
                            vendor_link = row.find_element(By.CSS_SELECTOR, "a[href*='fs.aspx']")
                            vendor_url = vendor_link.get_attribute('href')
                            button_text = vendor_link.text.strip() or "לפרטים נוספים"
                            print(f"   🔗 Found vendor URL (alt): {vendor_url[:100]}...")
                        except:
                            print(f"   ⚠️ Could not extract vendor URL from ModelRow")
                else:
                    # Original logic for comparison pages
                    try:
                        details_section = row.find_element(By.CSS_SELECTOR, ".compare-item-details")
                        vendor_link = details_section.find_element(By.CSS_SELECTOR, "a[href*='/fs']")
                        vendor_url = vendor_link.get_attribute('href')
                        button_text = vendor_link.text.strip() or "לפרטים נוספים"
                    except:
                        pass
                
                # UNIFIED: Extract individual vendor product name (visit vendor page)
                individual_vendor_product_name = extract_individual_vendor_product_name(driver, vendor_url, vendor_name)
                
                # Fallback: If vendor extraction failed, use ZAP product name with vendor indication
                if "Unknown Product" in individual_vendor_product_name:
                    zap_product_name = extract_zap_product_name(driver)
                    if zap_product_name and zap_product_name != "Unknown Product":
                        individual_vendor_product_name = f"{zap_product_name} (via {vendor_name})"
                
                # ENHANCED: Extract price based on page type
                vendor_price = 0
                
                if "models.aspx" in current_url:
                    # Extract price from listings page (models.aspx)
                    try:
                        # Look for price in ModelRow - multiple possible selectors
                        price_selectors = [
                            ".price-wrapper.product.final",  # Final price
                            ".price-wrapper.product.total",  # Total price  
                            "span[class*='price']",          # Any price span
                            ".ModelDetails span",            # Price in details
                        ]
                        
                        for selector in price_selectors:
                            try:
                                price_element = row.find_element(By.CSS_SELECTOR, selector)
                                price_text = price_element.text.strip()
                                print(f"   🔍 Price text from {selector}: '{price_text}'")
                                
                                if price_text and '₪' in price_text:
                                    # Extract number from Hebrew price format
                                    price_match = re.search(r'[\d,]+', price_text.replace('₪', ''))
                                    if price_match:
                                        price_str = price_match.group().replace(',', '')
                                        if price_str.isdigit() and len(price_str) >= 3:
                                            vendor_price = int(price_str)
                                            print(f"   ✅ Price extracted from models.aspx: ₪{vendor_price}")
                                            break
                            except:
                                continue
                                
                        if vendor_price == 0:
                            # Fallback: search for any text containing ₪ and numbers
                            row_text = row.text
                            price_matches = re.findall(r'₪?\s*[\d,]+\s*₪?', row_text)
                            for match in price_matches:
                                numbers = re.findall(r'\d+', match.replace(',', ''))
                                if numbers:
                                    price_candidate = int(''.join(numbers))
                                    if 100 <= price_candidate <= 50000:  # Reasonable price range
                                        vendor_price = price_candidate
                                        print(f"   ✅ Price extracted via text search: ₪{vendor_price}")
                                        break
                    except Exception as e:
                        print(f"   ❌ Models.aspx price extraction failed: {e}")
                else:
                    # Original logic for comparison pages
                    try:
                        # Use the WORKING selector proven by test: ".compare-item-details span"
                        price_element = row.find_element(By.CSS_SELECTOR, ".compare-item-details span")
                        price_text = price_element.text.strip()
                        print(f"   🔍 Price text: '{price_text}'")
                        
                        if price_text:
                            price_match = re.search(r'[\d,]+', price_text)
                            if price_match:
                                price_str = price_match.group().replace(',', '')
                                if price_str.isdigit() and len(price_str) >= 3:
                                    vendor_price = int(price_str)
                                    print(f"   ✅ Price extracted: ₪{vendor_price}")
                    
                    except Exception as e:
                        print(f"   ❌ Price extraction failed: {e}")
                
                vendors.append({
                    'vendor_name': vendor_name,
                    'vendor_price': vendor_price,
                    'vendor_url': vendor_url,
                    'button_text': button_text,
                    'vendor_product_name': individual_vendor_product_name
                })
                
                print(f"  {i}. {vendor_name} - ₪{vendor_price}")
                
            except Exception as e:
                print(f"  Error extracting vendor {i}: {e}")
                continue
        
        return vendors
        
    except Exception as e:
        print(f"❌ Unified vendor extraction failed: {e}")
        return []

def process_single_product(product_name, original_price, line_number):
    """Process a single product with unified approach."""
    driver = create_driver()
    start_time = time.time()
    
    try:
        print(f"\n{'='*70}")
        print(f"🚀 PROCESSING LINE {line_number}")
        print(f"📋 Product: {product_name}")
        print(f"💰 Original Price: ₪{original_price}")
        print(f"🤖 Mode: {'HEADLESS' if HEADLESS_MODE else 'VISIBLE'}")
        print(f"{'='*70}")
        
        # Step 1: Search and navigate to product page
        search_result, model_url = search_and_filter_product(driver, product_name)
        
        if search_result != "success":
            print(f"❌ Search failed: {search_result}")
            return {
                'line_number': line_number,
                'product_name': product_name,
                'original_price': original_price,
                'vendors': [],
                'search_result': search_result,
                'processing_time': time.time() - start_time,
                'model_url': model_url
            }
        
        # Step 2: Extract ZAP product name from current page
        zap_product_name = extract_zap_product_name(driver)
        
        # Step 3: Extract vendors using unified logic
        vendors = extract_vendors_unified(driver, zap_product_name)
        
        processing_time = time.time() - start_time
        
        print(f"\n📊 PROCESSING SUMMARY:")
        print(f"   Line: {line_number}")
        print(f"   Product: {product_name}")
        print(f"   Vendors found: {len(vendors)}")
        print(f"   Processing time: {processing_time:.1f}s")
        print(f"   Mode: {'HEADLESS' if HEADLESS_MODE else 'VISIBLE'}")
        
        # Keep browser open briefly for verification (shorter in headless)
        if not HEADLESS_MODE:
            print(f"🔍 Browser open for 3 seconds for verification...")
            time.sleep(3)
        
        return {
            'line_number': line_number,
            'product_name': product_name,
            'original_price': original_price,
            'vendors': vendors,
            'search_result': search_result,
            'processing_time': processing_time,
            'model_url': model_url
        }
        
    finally:
        driver.quit()

def create_excel_file(results, filename=None):
    """Create Excel file with CORRECT Hebrew format matching target_writer structure."""
    if not filename:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"output/Lines_{'-'.join(str(r['line_number']) for r in results)}_Report_{timestamp}.xlsx"
    
    # Create directories
    os.makedirs("output", exist_ok=True)
    
    # Create workbook with two sheets
    wb = openpyxl.Workbook()
    
    # Sheet 1: Details (פירוט) - CORRECT Hebrew headers
    details_sheet = wb.active
    details_sheet.title = "פירוט"
    
    # Headers matching the expected format exactly
    details_headers = [
        "שורת מקור",           # A - Source Row
        "שם מוצר",             # B - Product Name  
        "מחיר",                # C - Original Price
        "שם ספק",              # D - Vendor Name
        "שם מוצר באתר הספק",    # E - Product on Vendor Site
        "מחיר ZAP",            # F - ZAP Price (vendor price)
        "הפרש מחיר",           # G - Price Difference
        "% הפרש",              # H - Percentage Difference  
        "טקסט הכפתור",         # I - Button Text
        "קישור",               # J - Link
        "זמן עדכון",           # K - Update Timestamp
        "Model ID",            # L - Model ID
        "Method Used"          # M - Method Used
    ]
    
    # Style headers
    from openpyxl.styles import Font, PatternFill, Alignment
    for col, header in enumerate(details_headers, 1):
        cell = details_sheet.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Sheet 2: Summary (סיכום) - CORRECT Hebrew headers
    summary_sheet = wb.create_sheet("סיכום")
    
    # Sheet 3: Exceptions (חריגים) - NEW for rejected vendors
    exceptions_sheet = wb.create_sheet("חריגים")
    
    # Exceptions sheet headers
    exceptions_headers = [
        "מספר שורה מקור",      # A - Source Row Number
        "שם מוצר מקורי",       # B - Original Product Name
        "מחיר רשמי",           # C - Official Price
        "שם ספק",              # D - Vendor Name
        "שם מוצר באתר ספק",    # E - Product Name on Vendor Site
        "מחיר ספק",            # F - Vendor Price
        "הפרש מחיר",           # G - Price Difference
        "אחוז הפרש",           # H - Percentage Difference
        "קישור לספק",          # I - Vendor Link
        "חותמת זמן",           # J - Timestamp
        "ציון אימות",          # K - Validation Score
        "סטטוס",               # L - Status
        "סיבות דחייה"          # M - Rejection Reasons
    ]
    
    # Style exceptions headers
    for col, header in enumerate(exceptions_headers, 1):
        cell = exceptions_sheet.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="C65911", end_color="C65911", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    summary_headers = [
        "שורת מקור",           # A - Source Row
        "שם המוצר",            # B - Product Name
        "Model ID",            # C - Model ID  
        "מחיר מקורי",          # D - Original Price
        "סה\"כ ספקים",         # E - Total Vendors
        "מחיר מינימלי",        # F - Min Price
        "מחיר מקסימלי",        # G - Max Price
        "מחיר ממוצע",          # H - Average Price
        "חיסכון מקסימלי",      # I - Maximum Savings
        "שיטת חיפוש",          # J - Search Method
        "זמן עיבוד"            # K - Processing Time
    ]
    
    # Style summary headers
    for col, header in enumerate(summary_headers, 1):
        cell = summary_sheet.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Fill data
    for result in results:
        line_num = result['line_number']
        product_name = result['product_name']
        original_price = result['original_price']
        vendors = result['vendors']
        processing_time = result['processing_time']
        model_url = result.get('model_url', '')
        
        # Extract Model ID from URL
        model_id = ""
        if 'modelid=' in model_url:
            import re
            match = re.search(r'modelid=(\d+)', model_url)
            if match:
                model_id = match.group(1)
        
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        if vendors:
            # Process vendors and identify exceptions using REAL scoring engine
            accepted_vendors = []
            rejected_vendors = []
            
            # Initialize scoring engine
            scoring_engine = ProductScoringEngine()
            
            for vendor in vendors:
                vendor_product_name = vendor.get('vendor_product_name', '')
                
                # Use the REAL scoring engine to evaluate vendor product match
                if vendor_product_name and vendor_product_name != 'Unknown Product':
                    # Calculate score using the scoring engine
                    scoring_result = scoring_engine.calculate_match_score(
                        product_name,  # Original product name
                        vendor_product_name  # Vendor's product name
                    )
                    
                    validation_score = scoring_result.total_score
                    vendor['validation_score'] = validation_score
                    
                    # Collect rejection reasons from scoring issues
                    rejection_reasons = []
                    
                    # Check gates
                    if not scoring_result.gates_passed.get('model_number'):
                        rejection_reasons.append("Model number gate failed")
                    if not scoring_result.gates_passed.get('product_type'):
                        rejection_reasons.append("Product type gate failed")
                    
                    # Add other issues from scoring
                    rejection_reasons.extend(scoring_result.issues)
                    
                    # Add extreme price difference check
                    if vendor['vendor_price'] > 0 and original_price > 0:
                        price_diff_pct = abs((vendor['vendor_price'] - original_price) / original_price * 100)
                        if price_diff_pct > 50:
                            rejection_reasons.append(f"Extreme price difference: {price_diff_pct:.1f}%")
                    
                    vendor['rejection_reasons'] = "; ".join(rejection_reasons) if rejection_reasons else ""
                    
                    # Apply threshold (8.0/10.0 as per CLAUDE.md)
                    if validation_score < 8.0:
                        rejected_vendors.append(vendor)
                    else:
                        accepted_vendors.append(vendor)
                else:
                    # No vendor product name available
                    vendor['validation_score'] = 0.0
                    vendor['rejection_reasons'] = "No product name from vendor site"
                    rejected_vendors.append(vendor)
            
            # Details sheet - one row per ACCEPTED vendor with PROPER FORMATTING
            for i, vendor in enumerate(accepted_vendors, 2):  # Start from row 2
                if vendor['vendor_price'] > 0:
                    price_diff = vendor['vendor_price'] - original_price
                    price_diff_pct = (price_diff / original_price * 100) if original_price > 0 else 0
                    
                    # Format prices with ₪ symbol
                    formatted_original = f"₪{original_price:,.1f}"
                    formatted_vendor = f"₪{vendor['vendor_price']:,}"
                    formatted_diff = f"₪{price_diff:,.1f}"
                    
                    details_sheet.append([
                        str(line_num),              # שורת מקור
                        product_name,               # שם מוצר
                        formatted_original,         # מחיר
                        vendor['vendor_name'],      # שם ספק
                        vendor.get('vendor_product_name', 'Unknown Product'),  # שם מוצר באתר הספק
                        formatted_vendor,           # מחיר ZAP
                        formatted_diff,             # הפרש מחיר
                        f"{price_diff_pct:.1f}%",   # % הפרש
                        vendor.get('button_text', 'לפרטים נוספים'),  # טקסט הכפתור
                        vendor['vendor_url'],       # קישור
                        timestamp,                  # זמן עדכון
                        model_id,                   # Model ID
                        "Breakthrough - Dropdown"   # Method Used
                    ])
            
            # Exceptions sheet - one row per REJECTED vendor
            for vendor in rejected_vendors:
                if vendor['vendor_price'] > 0:
                    price_diff = vendor['vendor_price'] - original_price
                    price_diff_pct = (price_diff / original_price * 100) if original_price > 0 else 0
                    
                    exceptions_sheet.append([
                        str(line_num),                                        # מספר שורה מקור
                        product_name,                                         # שם מוצר מקורי
                        f"₪{original_price:,.1f}",                          # מחיר רשמי
                        vendor['vendor_name'],                               # שם ספק
                        vendor.get('vendor_product_name', 'Unknown'),        # שם מוצר באתר ספק
                        f"₪{vendor['vendor_price']:,}",                     # מחיר ספק
                        f"₪{price_diff:,.1f}",                              # הפרש מחיר
                        f"{price_diff_pct:.1f}%",                           # אחוז הפרש
                        vendor.get('vendor_url', ''),                        # קישור לספק
                        timestamp,                                            # חותמת זמן
                        f"{vendor.get('validation_score', 0.0):.1f}/10.0",  # ציון אימות
                        "⚠️ דורש בדיקה",                                    # סטטוס
                        vendor.get('rejection_reasons', '')                  # סיבות דחייה
                    ])
            
            # Summary sheet - one row per product with PROPER FORMATTING
            valid_prices = [v['vendor_price'] for v in accepted_vendors if v['vendor_price'] > 0]
            if valid_prices:
                min_price = min(valid_prices)
                max_price = max(valid_prices)
                avg_price = sum(valid_prices) / len(valid_prices)
                max_savings = original_price - min_price
            else:
                min_price = max_price = avg_price = max_savings = 0
            
            summary_sheet.append([
                str(line_num),                  # שורת מקור
                product_name,                   # שם המוצר
                model_id,                       # Model ID
                f"₪{original_price:,.1f}",     # מחיר מקורי
                str(len(accepted_vendors)),    # סה"כ ספקים (accepted only)
                f"₪{min_price:,}",             # מחיר מינימלי
                f"₪{max_price:,}",             # מחיר מקסימלי
                f"₪{avg_price:,.0f}",          # מחיר ממוצע
                f"₪{max_savings:,.1f}",        # חיסכון מקסימלי
                "Breakthrough - Dropdown",      # שיטת חיפוש
                f"{processing_time:.1f}s"       # זמן עיבוד
            ])
        else:
            # No vendors found
            summary_sheet.append([
                str(line_num), product_name, model_id, f"₪{original_price:,.1f}", 
                "0", "₪0", "₪0", "₪0", "₪0", "Failed", f"{processing_time:.1f}s"
            ])
    
    # Auto-adjust column widths
    for sheet in [details_sheet, summary_sheet, exceptions_sheet]:
        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width
    
    # Save file
    wb.save(filename)
    print(f"📊 Excel file saved: {filename}")
    return filename

def main():
    """Main function with unified processing."""
    global HEADLESS_MODE
    
    parser = argparse.ArgumentParser(description='Production Universal Product Scraper - UNIFIED VERSION')
    parser.add_argument('--headless', action='store_true', help='Run in headless mode')
    parser.add_argument('lines', nargs='+', type=int, help='Line numbers to process')
    
    args = parser.parse_args()
    
    HEADLESS_MODE = args.headless
    target_lines = args.lines
    
    print("🚀 UNIFIED PRODUCTION SCRAPER")
    print("="*70)
    print(f"📋 Processing Lines: {target_lines}")
    print(f"🤖 Mode: {'HEADLESS' if HEADLESS_MODE else 'VISIBLE'}")
    print("="*70)
    
    # Read source data
    try:
        reader = SourceExcelReader(start_row=2)
        products = reader.read_products('data/SOURCE.xlsx')
        print(f"✅ Loaded {len(products)} products from source")
    except Exception as e:
        print(f"❌ Failed to read source data: {e}")
        return
    
    # Process specified lines - NOW USING EXCEL ROW NUMBERS DIRECTLY
    results = []
    for line_num in target_lines:
        # Find product with matching Excel row number
        product_found = None
        for product in products:
            if product.row_number == line_num:
                product_found = product
                break
        
        if product_found:
            result = process_single_product(
                product_found.name, 
                product_found.original_price, 
                line_num
            )
            results.append(result)
        else:
            print(f"❌ Excel row {line_num} not found in source data (might be empty or header row)")
    
    # Create Excel output using proper TargetExcelWriter
    if results:
        # Convert results to ProductScrapingResult format for TargetExcelWriter
        from src.excel.target_writer import TargetExcelWriter
        from src.models.data_models import ProductScrapingResult, ProductInput, VendorOffer
        
        # Convert results to proper format
        formatted_results = []
        for result in results:
            # Create ProductInput - parse the product name into components
            product_name_parts = result['product_name'].split(' ')
            manufacturer = product_name_parts[0] if product_name_parts else ''
            model_series = ' '.join(product_name_parts[1:-1]) if len(product_name_parts) > 2 else ''
            model_number = product_name_parts[-1] if len(product_name_parts) > 1 else ''
            
            product_input = ProductInput(
                row_number=result['line_number'],
                manufacturer=manufacturer,
                model_series=model_series,
                model_number=model_number,
                original_price=result['original_price']
            )
            
            # Create VendorOffers
            vendor_offers = []
            for vendor in result['vendors']:
                offer = VendorOffer(
                    vendor_name=vendor['vendor_name'],
                    product_name=vendor['vendor_product_name'],
                    price=vendor['vendor_price'],
                    url=vendor['vendor_url'],
                    button_text=vendor.get('button_text', 'לפרטים נוספים')
                )
                vendor_offers.append(offer)
            
            # Create ProductScrapingResult
            scraping_result = ProductScrapingResult(
                input_product=product_input,
                vendor_offers=vendor_offers,
                status="success" if vendor_offers else "no_results",
                error_message=None,
                model_id=result.get('model_id', ''),
                listing_count=len(vendor_offers),
                constructed_url=result.get('model_url', '')
            )
            formatted_results.append(scraping_result)
        
        # Use proper Excel writer
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_file = f"output/Lines_{'-'.join(str(r['line_number']) for r in results)}_Report_{timestamp}.xlsx"
        
        excel_writer = TargetExcelWriter()
        success = excel_writer.write_results(formatted_results, excel_file)
        
        if not success:
            print("❌ Failed to write Excel file")
            excel_file = None
        
        print(f"\n🎯 FINAL SUMMARY:")
        print(f"   Lines processed: {len(results)}")
        print(f"   Mode: {'HEADLESS' if HEADLESS_MODE else 'VISIBLE'}")
        print(f"   Excel file: {excel_file}")
        
        # Show key statistics
        total_vendors = sum(len(r['vendors']) for r in results)
        avg_processing_time = sum(r['processing_time'] for r in results) / len(results)
        
        print(f"   Total vendors: {total_vendors}")
        print(f"   Avg processing time: {avg_processing_time:.1f}s")
        
        return excel_file
    
    return None

if __name__ == "__main__":
    main()
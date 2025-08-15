"""
Excel source file reader for Universal Product Scraper.

Reads product information from SOURCE Excel files.
"""

import os
from typing import List, Optional
from pathlib import Path

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from src.models.data_models import ProductInput
from src.utils.logger import get_logger
from src.utils.exceptions import ExcelException, ValidationException


logger = get_logger(__name__)


class SourceExcelReader:
    """Read product data from SOURCE Excel files."""
    
    def __init__(self, start_row: int = 2):
        """
        Initialize reader with configuration for new Hebrew structure.
        
        Args:
            start_row: Row number where data starts (1-indexed) - starts at 2 due to Hebrew headers
        """
        self.start_row = start_row
        
        # NEW Column mappings for Hebrew structure (1-indexed)
        # Headers: יצרן / יבואן | דגם | מספר | מחיר
        self.col_manufacturer = 1    # Column A: יצרן / יבואן  
        self.col_model_series = 2    # Column B: דגם
        self.col_model_number = 3    # Column C: מספר
        self.col_original_price = 4  # Column D: מחיר
    
    def read_products(self, file_path: str) -> List[ProductInput]:
        """
        Read products from Excel file.
        
        Args:
            file_path: Path to SOURCE Excel file
            
        Returns:
            List of ProductInput objects
            
        Raises:
            ExcelException: If file cannot be read
            ValidationException: If file format is invalid
        """
        if not os.path.exists(file_path):
            raise ExcelException(f"Source file not found: {file_path}")
        
        logger.info(f"Reading products from: {file_path}")
        
        try:
            # Load workbook
            workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            
            # Get first worksheet
            if not workbook.worksheets:
                raise ValidationException("No worksheets found in Excel file")
            
            worksheet = workbook.active
            logger.info(f"Reading from worksheet: {worksheet.title}")
            
            # Read products
            products = self._read_products_from_worksheet(worksheet)
            
            # Close workbook
            workbook.close()
            
            logger.info(f"Successfully read {len(products)} products")
            return products
            
        except openpyxl.utils.exceptions.InvalidFileException as e:
            raise ExcelException(f"Invalid Excel file format: {e}")
        except Exception as e:
            raise ExcelException(f"Failed to read Excel file: {e}")
    
    def _read_products_from_worksheet(self, worksheet: Worksheet) -> List[ProductInput]:
        """Read products from a worksheet."""
        products = []
        empty_rows = 0
        max_empty_rows = 5  # Stop after 5 consecutive empty rows
        
        # Start reading from the configured row
        for row_idx in range(self.start_row, worksheet.max_row + 1):
            try:
                # Get cell values from NEW Hebrew structure
                manufacturer = worksheet.cell(row=row_idx, column=self.col_manufacturer).value
                model_series = worksheet.cell(row=row_idx, column=self.col_model_series).value  
                model_number = worksheet.cell(row=row_idx, column=self.col_model_number).value
                original_price = worksheet.cell(row=row_idx, column=self.col_original_price).value
                
                # Check if row is empty (any of the main components missing)
                if not manufacturer or not model_series or not model_number:
                    empty_rows += 1
                    if empty_rows >= max_empty_rows:
                        logger.info(f"Stopping at row {row_idx} after {max_empty_rows} empty rows")
                        break
                    continue
                
                # Reset empty row counter
                empty_rows = 0
                
                # Validate and clean component data
                manufacturer = str(manufacturer).strip()
                model_series = str(model_series).strip()
                model_number = str(model_number).strip()
                
                # Skip header rows (containing Hebrew header terms)
                header_terms = [
                    'יצרן', 'יבואן', 'דגם', 'מספר', 'מחיר', 'manufacturer', 'importer', 'model', 'number', 'price'
                ]
                if any(term in manufacturer.lower() or term in model_series.lower() for term in header_terms):
                    logger.debug(f"Skipping header row {row_idx}: {manufacturer} {model_series}")
                    continue
                
                # Skip rows with invalid components (too short, etc.)
                if len(manufacturer) < 2 or len(model_series) < 2 or len(model_number) < 1:
                    logger.debug(f"Skipping invalid components at row {row_idx}: {manufacturer} | {model_series} | {model_number}")
                    continue
                
                # CRITICAL: Skip rows without valid price (as per user requirement)
                # All components AND price must exist for a valid product row
                if original_price is None or original_price == "":
                    combined_name = f"{manufacturer} {model_series} {model_number}"
                    logger.debug(f"Skipping row {row_idx} - no price: '{combined_name}'")
                    continue
                
                # Try to convert price to number
                try:
                    price_value = float(original_price)
                    if price_value <= 0:
                        combined_name = f"{manufacturer} {model_series} {model_number}"
                        logger.debug(f"Skipping row {row_idx} - invalid price ({original_price}): '{combined_name}'")
                        continue
                except (ValueError, TypeError):
                    combined_name = f"{manufacturer} {model_series} {model_number}"
                    logger.debug(f"Skipping row {row_idx} - non-numeric price ({original_price}): '{combined_name}'")
                    continue
                
                # Handle price (we already validated it exists and is numeric above)
                original_price = price_value
                
                # Create ProductInput with NEW component-based structure
                product = ProductInput(
                    row_number=row_idx,  # Use row index as identifier
                    manufacturer=manufacturer,
                    model_series=model_series,
                    model_number=model_number,
                    original_price=original_price
                )
                
                products.append(product)
                logger.debug(f"Read product: {product}")
                
            except Exception as e:
                logger.error(f"Error reading row {row_idx}: {e}")
                continue
        
        return products
    
    def validate_format(self, file_path: str) -> bool:
        """
        Validate Excel file format.
        
        Args:
            file_path: Path to Excel file
            
        Returns:
            True if format is valid
            
        Raises:
            ExcelException: If file cannot be opened
        """
        try:
            workbook = openpyxl.load_workbook(file_path, read_only=True)
            worksheet = workbook.active
            
            # Check if we have data at the start row (NEW structure validation)
            has_data = False
            for row_idx in range(self.start_row, min(self.start_row + 10, worksheet.max_row + 1)):
                manufacturer = worksheet.cell(row=row_idx, column=self.col_manufacturer).value
                model_series = worksheet.cell(row=row_idx, column=self.col_model_series).value
                if manufacturer and model_series:
                    has_data = True
                    break
            
            workbook.close()
            
            if not has_data:
                logger.warning(f"No product data found starting from row {self.start_row}")
                return False
            
            return True
            
        except Exception as e:
            logger.error(f"Failed to validate Excel format: {e}")
            return False
    
    def get_preview(self, file_path: str, num_rows: int = 5) -> List[dict]:
        """
        Get a preview of the first few products.
        
        Args:
            file_path: Path to Excel file
            num_rows: Number of rows to preview
            
        Returns:
            List of dictionaries with product data
        """
        try:
            products = self.read_products(file_path)[:num_rows]
            return [
                {
                    "row": p.row_number,
                    "manufacturer": p.manufacturer,
                    "model_series": p.model_series,
                    "model_number": p.model_number,
                    "combined_name": p.name,
                    "price": p.original_price,
                    "search_term": p.search_term
                }
                for p in products
            ]
        except Exception as e:
            logger.error(f"Failed to get preview: {e}")
            return [] 
"""
Excel target file writer for Universal Product Scraper.

Writes scraping results to TARGET Excel files with 2 worksheets.
"""

import os
from datetime import datetime
from typing import List, Tuple
import statistics

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from src.models.data_models import ProductScrapingResult, ProductSummary, VendorOffer
from src.utils.logger import get_logger
from src.utils.exceptions import ExcelException


logger = get_logger(__name__)


class TargetExcelWriter:
    """Write scraping results to TARGET Excel with 2 worksheets."""
    
    def __init__(self):
        """Initialize writer with Hebrew column headers."""
        # Worksheet 1 (Details) headers
        self.details_headers = [
            "שורת מקור",          # A - Source Row
            "שם מוצר",            # B - Product Name
            "מחיר",               # C - Original Price
            "שם ספק",            # D - Vendor Name
            "שם מוצר באתר הספק", # E - Product Name on Vendor Site
            "מחיר ZAP",           # F - ZAP Price (vendor price)
            "הפרש מחיר",          # G - Price Difference
            "% הפרש",             # H - Percentage Difference
            "טקסט הכפתור",        # I - Button Text (קנו עכשיו, לפרטים נוספים, השוואת מחירים)
            "קישור",              # J - Link
            "זמן עדכון"           # K - Update Timestamp
        ]
        
        # Worksheet 2 (Summary) headers - UNIFIED FORMAT (17 columns with dual URLs)
        self.summary_headers_unified = [
            "שם מוצר",           # A - Product Name
            "Model ID",          # B - ZAP Model ID
            "רשומות שנמצאו",     # C - Listings Found - Total found after filtering
            "רשומות שנגרדו",     # D - Listings Processed - Actually scraped
            "סטטוס",             # E - Status
            "מחיר מינימלי",      # F - Minimum Price
            "מחיר מקסימלי",      # G - Maximum Price
            "מחיר ממוצע",        # H - Average Price
            "טווח מחירים",       # I - Price Range
            "מספר ספקים",        # J - Number of Vendors
            "ספק הזול ביותר",    # K - Cheapest Vendor Name
            "קישור ל Opt.1",     # L - Option 1 URL (Model ID URL)
            "סטיית תקן",         # M - Standard Deviation
            "מחיר מקורי",        # N - Original Reference Price
            "% הפרש ממקור",      # O - % Difference from Original
            "זמן עדכון"          # P - Update Timestamp
        ]
        
        # Keep the old headers for backward compatibility (but use unified format)
        self.summary_headers_success = self.summary_headers_unified
        
        # Use the unified format for failure cases too
        self.summary_headers_failure = self.summary_headers_unified
    
    def write_failure_results(self, product_input, dual_approach_result, target_path: str) -> bool:
        """
        Write FAILURE results to Excel with Summary sheet only (no Details sheet).
        
        Args:
            product_input: ProductInput object
            dual_approach_result: Result from dual_approach_validation (failed)
            target_path: Output Excel file path
            
        Returns:
            True if successful
        """
        try:
            from datetime import datetime
            
            # Create failure summary
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            failure_summary = self.create_failure_summary(product_input, dual_approach_result, timestamp)
            
            # Create workbook with only summary sheet
            workbook = Workbook()
            
            # Remove default sheet
            default_sheet = workbook.active
            workbook.remove(default_sheet)
            
            # Create failure summary worksheet
            self.create_summary_worksheet_failure(workbook, [failure_summary])
            
            # Save workbook
            workbook.save(target_path)
            
            logger.info(f"✅ Failure results written to: {target_path}")
            logger.info(f"📋 Summary: FAILURE - {failure_summary['status']}")
            logger.info(f"🔗 Option 1 URL: {failure_summary['option1_url']}")
            
            return True
            
        except Exception as e:
            logger.error(f"❌ Failed to write failure results to {target_path}: {e}")
            return False

    def write_results(self, results: List[ProductScrapingResult], target_path: str) -> bool:
        """
        Write SUCCESS results to Excel with Details and Summary sheets.
        
        Args:
            results: List of scraping results
            target_path: Output Excel file path
            
        Returns:
            True if successful
            
        Raises:
            ExcelException: If write fails
        """
        try:
            # Create directory if needed
            target_dir = os.path.dirname(target_path)
            if target_dir:
                os.makedirs(target_dir, exist_ok=True)
            
            logger.info(f"Writing results to: {target_path}")
            
            # Create workbook
            workbook = Workbook()
            
            # Create worksheet 1: Details
            self.create_detailed_worksheet(workbook, results)
            
            # Create worksheet 2: Summary (if any products have results)
            if self._validate_all_products_scraped(results):
                self.create_summary_worksheet(workbook, results)
            else:
                logger.warning("No products scraped successfully, skipping summary sheet")
                # Still create empty summary sheet for consistency
                self._create_empty_summary_sheet(workbook)
            
            # Save file
            workbook.save(target_path)
            workbook.close()
            
            logger.info(f"Successfully wrote results to {target_path}")
            return True
            
        except Exception as e:
            logger.error(f"Failed to write Excel file: {e}")
            raise ExcelException(f"Excel write failed: {e}")
    
    def create_detailed_worksheet(self, workbook: Workbook, results: List[ProductScrapingResult]) -> None:
        """Create the detailed worksheet with all vendor offers."""
        # Use the active sheet and rename it
        ws = workbook.active
        ws.title = "פירוט"
        
        # Write headers
        for col, header in enumerate(self.details_headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            self._style_header_cell(cell)
        
        # Write data
        row_num = 2
        for result in results:
            if not result.vendor_offers:
                # Write a row even if no vendors found
                ws.cell(row=row_num, column=1, value=result.input_product.row_number)
                ws.cell(row=row_num, column=2, value=result.input_product.name)
                ws.cell(row=row_num, column=3, value=result.input_product.original_price)
                ws.cell(row=row_num, column=4, value="לא נמצאו ספקים")
                ws.cell(row=row_num, column=11, value=datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
                row_num += 1
            else:
                # Write a row for each vendor offer
                for offer in result.vendor_offers:
                    # Calculate price differences
                    price_diff, price_diff_pct = self.calculate_price_difference(
                        result.input_product.original_price,
                        offer.price
                    )
                    
                    # Write row data
                    ws.cell(row=row_num, column=1, value=result.input_product.row_number)
                    ws.cell(row=row_num, column=2, value=result.input_product.name)
                    ws.cell(row=row_num, column=3, value=result.input_product.original_price)
                    ws.cell(row=row_num, column=4, value=offer.vendor_name)
                    ws.cell(row=row_num, column=5, value=offer.product_name)
                    ws.cell(row=row_num, column=6, value=offer.price)
                    ws.cell(row=row_num, column=7, value=price_diff)
                    ws.cell(row=row_num, column=8, value=price_diff_pct / 100)  # Convert percentage to ratio for Excel formatting
                    ws.cell(row=row_num, column=9, value=offer.button_text)  # Button text column
                    ws.cell(row=row_num, column=10, value=offer.url)  # URL moved to column 10
                    ws.cell(row=row_num, column=11, value=datetime.now().strftime('%Y-%m-%d %H:%M:%S'))  # Timestamp moved to column 11
                    
                    # Make URL a hyperlink (now in column 10)
                    if offer.url:
                        ws.cell(row=row_num, column=10).hyperlink = offer.url
                        ws.cell(row=row_num, column=10).style = "Hyperlink"
                    
                    row_num += 1
        
        # Format columns with FIXED formatting
        self.format_excel_columns_fixed(ws)
    
    def create_summary_worksheet_success(self, workbook: Workbook, results: List[ProductScrapingResult]) -> None:
        """Create the summary worksheet for SUCCESS cases with statistics."""
        ws = workbook.create_sheet("סיכום")
        
        # Write SUCCESS headers
        for col, header in enumerate(self.summary_headers_success, 1):
            cell = ws.cell(row=1, column=col, value=header)
            self._style_header_cell(cell)
        
        # Calculate and write summary data for successful results
        summaries = self.calculate_summary_statistics(results)
        
        for row_num, summary in enumerate(summaries, 2):
            ws.cell(row=row_num, column=1, value=summary.product_name)
            ws.cell(row=row_num, column=2, value=summary.model_id)                    # Model ID
            ws.cell(row=row_num, column=3, value=summary.listings_found)             # Total Found
            ws.cell(row=row_num, column=4, value=summary.listings_processed)         # Actually Processed
            ws.cell(row=row_num, column=5, value=summary.status)                     # Status
            ws.cell(row=row_num, column=6, value=summary.minimum_price)              # Min Price
            ws.cell(row=row_num, column=7, value=summary.maximum_price)              # Max Price
            ws.cell(row=row_num, column=8, value=round(summary.average_price, 2))    # Avg Price
            ws.cell(row=row_num, column=9, value=summary.price_range)                # Price Range
            ws.cell(row=row_num, column=10, value=summary.vendor_count)              # Vendor Count
            ws.cell(row=row_num, column=11, value=summary.cheapest_vendor_name)      # Cheapest Vendor
            ws.cell(row=row_num, column=12, value=getattr(summary, 'option1_url', ''))  # Opt.1 URL
            ws.cell(row=row_num, column=13, value=round(summary.standard_deviation, 2)) # Std Dev
            ws.cell(row=row_num, column=14, value=summary.zap_reference_price)       # ZAP Price
            ws.cell(row=row_num, column=15, value=round(summary.percentage_diff_from_zap / 100, 4)) # % Diff
            ws.cell(row=row_num, column=16, value=summary.update_timestamp)          # Timestamp
            
            # Make Opt.1 URL hyperlink
            if getattr(summary, 'option1_url', ''):
                ws.cell(row=row_num, column=12).hyperlink = summary.option1_url
                ws.cell(row=row_num, column=12).style = "Hyperlink"
        
        # Format columns with FIXED formatting
        self.format_excel_columns_fixed(ws)

    def create_summary_worksheet_failure(self, workbook: Workbook, failure_summaries: List[dict]) -> None:
        """Create the summary worksheet for FAILURE cases with dual URLs."""
        ws = workbook.create_sheet("סיכום")
        
        # Write FAILURE headers
        for col, header in enumerate(self.summary_headers_failure, 1):
            cell = ws.cell(row=1, column=col, value=header)
            self._style_header_cell(cell)
        
        # Write failure summary data
        for row_num, summary in enumerate(failure_summaries, 2):
            ws.cell(row=row_num, column=1, value=summary['product_name'])
            ws.cell(row=row_num, column=2, value=summary['model_id'])               # Failure message
            ws.cell(row=row_num, column=3, value=summary['listings_found'])        # Empty
            ws.cell(row=row_num, column=4, value=summary['listings_processed'])    # Empty
            ws.cell(row=row_num, column=5, value=summary['status'])                # Failure message
            ws.cell(row=row_num, column=6, value=summary['minimum_price'])         # Empty
            ws.cell(row=row_num, column=7, value=summary['maximum_price'])         # Empty
            ws.cell(row=row_num, column=8, value=summary['average_price'])         # Empty
            ws.cell(row=row_num, column=9, value=summary['price_range'])           # Empty
            ws.cell(row=row_num, column=10, value=summary['vendor_count'])         # Empty
            ws.cell(row=row_num, column=11, value=summary['cheapest_vendor_name']) # Empty
            ws.cell(row=row_num, column=12, value=summary['option1_url'])          # Option 1 URL
            ws.cell(row=row_num, column=13, value=summary['standard_deviation'])   # Empty
            ws.cell(row=row_num, column=14, value=summary['zap_reference_price'])  # Original price
            ws.cell(row=row_num, column=15, value=summary['percentage_diff_from_zap']) # Empty
            ws.cell(row=row_num, column=16, value=summary['update_timestamp'])     # Timestamp
            
            # Make Option URL hyperlink
            if summary['option1_url']:
                ws.cell(row=row_num, column=12).hyperlink = summary['option1_url']
                ws.cell(row=row_num, column=12).style = "Hyperlink"
        
        # Format columns with FIXED formatting
        self.format_excel_columns_fixed(ws)

    def create_summary_worksheet(self, workbook: Workbook, results: List[ProductScrapingResult]) -> None:
        """Create the summary worksheet with statistics - LEGACY METHOD."""
        # This method is kept for backward compatibility
        self.create_summary_worksheet_success(workbook, results)
    
    def calculate_price_difference(self, original_price: float, vendor_price: float) -> Tuple[float, float]:
        """
        Calculate price difference and percentage.
        
        Args:
            original_price: Original/reference price
            vendor_price: Vendor's price
            
        Returns:
            Tuple of (absolute_difference, percentage_difference)
        """
        if original_price <= 0:
            return vendor_price, 0.0
        
        diff = vendor_price - original_price
        # Calculate percentage difference: ((vendor_price - original_price) / original_price) * 100
        pct = ((vendor_price - original_price) / original_price) * 100
        
        return round(diff, 2), round(pct, 2)
    
    def create_failure_summary(self, product_input, dual_approach_result, timestamp) -> dict:
        """
        Create failure summary for products where both options failed.
        
        Args:
            product_input: ProductInput object
            dual_approach_result: Result from dual_approach_validation
            timestamp: Current timestamp
            
        Returns:
            Dictionary with failure summary data
        """
        # Extract URLs from dual approach result
        option1_url = ""
        if 'option1_failed' in dual_approach_result:
            option1_url = dual_approach_result['option1_failed'].get('url', '')
        
        return {
            'product_name': product_input.name,
            'model_id': "No Model ID fits the name of the product",
            'listings_found': None,
            'listings_processed': None,
            'status': "Failure - no matching products found",
            'minimum_price': None,
            'maximum_price': None,
            'average_price': None,
            'price_range': None,
            'vendor_count': None,
            'cheapest_vendor_name': None,
            'option1_url': option1_url,
            'standard_deviation': None,
            'zap_reference_price': product_input.original_price,
            'percentage_diff_from_zap': None,
            'update_timestamp': timestamp
        }

    def calculate_summary_statistics(self, results: List[ProductScrapingResult]) -> List[ProductSummary]:
        """Calculate summary statistics for each product - SUCCESS CASES ONLY."""
        summaries = []
        
        for result in results:
            if not result.vendor_offers:
                # Skip products with no vendors (failures handled separately)
                continue
            
            prices = [offer.price for offer in result.vendor_offers]
            
            # Find cheapest vendor
            cheapest_offer = min(result.vendor_offers, key=lambda x: x.price)
            
            # Calculate statistics
            std_dev = statistics.stdev(prices) if len(prices) > 1 else 0.0
            min_price = min(prices)
            max_price = max(prices)
            avg_price = sum(prices) / len(prices)
            
            # Calculate percentage as difference: ((min_price - original_price) / original_price) * 100  
            pct_diff = (((min_price - result.input_product.original_price) / result.input_product.original_price) * 100) if result.input_product.original_price > 0 else 0
            
            summary = ProductSummary(
                product_name=result.input_product.name,
                model_id=result.model_id or "N/A",                              # NEW: Model ID
                listings_found=result.listing_count or 0,                       # NEW: Total found
                listings_processed=len(result.vendor_offers),                   # NEW: Actually processed
                status=result.status,                                           # NEW: Status
                minimum_price=min_price,
                maximum_price=max_price,
                average_price=avg_price,
                price_range=max_price - min_price,
                vendor_count=len(result.vendor_offers),
                cheapest_vendor_name=cheapest_offer.vendor_name,
                option1_url=getattr(result, 'option1_url', ''),                 # NEW: Option 1 URL
                standard_deviation=std_dev,
                zap_reference_price=result.input_product.original_price,
                percentage_diff_from_zap=pct_diff,
                update_timestamp=datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            )
            summaries.append(summary)
        
        return summaries
    
    def format_excel_columns_fixed(self, worksheet) -> None:
        """Apply CORRECTED formatting to worksheet columns."""
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value:
                        # Consider Hebrew text might need more width
                        cell_length = len(str(cell.value)) * 1.2
                        if max_length < cell_length:
                            max_length = cell_length
                except:
                    pass
            
            # Set minimum and maximum widths
            adjusted_width = min(max(max_length, 10), 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Apply number formatting based on worksheet type
        sheet_name = worksheet.title
        for row in worksheet.iter_rows(min_row=2):
            
            if sheet_name == 'פירוט':  # Details sheet
                # Format prices (columns with price data) - Remove decimals, add shekel symbol
                for col_idx in [3, 6, 7]:  # Original price, vendor price, difference  
                    if col_idx <= len(row):
                        cell = row[col_idx - 1]
                        if isinstance(cell.value, (int, float)):
                            cell.number_format = '₪#,##0'  # Shekel symbol, no decimals
                
                # Format percentages
                for col_idx in [8]:  # Percentage columns in details
                    if col_idx <= len(row):
                        cell = row[col_idx - 1]
                        if isinstance(cell.value, (int, float)):
                            cell.number_format = '0.00%'
                            
            elif sheet_name == 'סיכום':  # Summary sheet  
                # Detect if this is failure case (17 columns) or success case (16 columns)
                total_columns = len(row)
                is_failure_case = total_columns >= 17
                
                # C,D = Counter fields (NO currency) - same for both cases
                for col_idx in [3, 4]:  # C=רשומות שנמצאו, D=רשומות שנגרדו
                    if col_idx <= len(row):
                        cell = row[col_idx - 1]
                        if isinstance(cell.value, (int, float)):
                            cell.number_format = '#,##0'  # Plain number with thousands separator
                
                if is_failure_case:
                    # FAILURE CASE: 16 columns with Opt.1 URL only
                    # F,G,H,I,N = Currency fields (WITH shekel symbol)
                    for col_idx in [6, 7, 8, 9, 14]:  # F=Min, G=Max, H=Avg, I=Range, N=ZAP price
                        if col_idx <= len(row):
                            cell = row[col_idx - 1]
                            if isinstance(cell.value, (int, float)):
                                cell.number_format = '₪#,##0'  # Shekel symbol with thousands separator
                    
                    # J = Vendor Count (NO currency)
                    for col_idx in [10]:  # J=מספר ספקים
                        if col_idx <= len(row):
                            cell = row[col_idx - 1]
                            if isinstance(cell.value, (int, float)):
                                cell.number_format = '#,##0'  # Plain number with thousands separator
                    
                    # N = Standard Deviation/Variance (NO currency but with 2 decimals)
                    for col_idx in [14]:  # N=סטיית תקן (shifted)
                        if col_idx <= len(row):
                            cell = row[col_idx - 1]
                            if isinstance(cell.value, (int, float)):
                                cell.number_format = '#,##0.00'  # Number with thousands separator and 2 decimals
                    
                    # P = Percentage field
                    for col_idx in [16]:  # P=% הפרש מ-ZAP (shifted)
                        if col_idx <= len(row):
                            cell = row[col_idx - 1]
                            if isinstance(cell.value, (int, float)):
                                cell.number_format = '0.00%'  # Percentage format
                else:
                    # SUCCESS CASE: 16 columns with single vendor URL
                    # F,G,H,I,N = Currency fields (WITH shekel symbol)
                    for col_idx in [6, 7, 8, 9, 14]:  # F=Min, G=Max, H=Avg, I=Range, N=ZAP price
                        if col_idx <= len(row):
                            cell = row[col_idx - 1]
                            if isinstance(cell.value, (int, float)):
                                cell.number_format = '₪#,##0'  # Shekel symbol with thousands separator
                    
                    # J = Vendor Count (NO currency)
                    for col_idx in [10]:  # J=מספר ספקים
                        if col_idx <= len(row):
                            cell = row[col_idx - 1]
                            if isinstance(cell.value, (int, float)):
                                cell.number_format = '#,##0'  # Plain number with thousands separator
                    
                    # M = Standard Deviation/Variance (NO currency but with 2 decimals)
                    for col_idx in [13]:  # M=סטיית תקן
                        if col_idx <= len(row):
                            cell = row[col_idx - 1]
                            if isinstance(cell.value, (int, float)):
                                cell.number_format = '#,##0.00'  # Number with thousands separator and 2 decimals
                    
                    # O = Percentage field
                    for col_idx in [15]:  # O=% הפרש מ-ZAP
                        if col_idx <= len(row):
                            cell = row[col_idx - 1]
                            if isinstance(cell.value, (int, float)):
                                cell.number_format = '0.00%'  # Percentage format
    
    def _style_header_cell(self, cell) -> None:
        """Apply styling to header cells."""
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def _validate_all_products_scraped(self, results: List[ProductScrapingResult]) -> bool:
        """Check if all products have been successfully scraped."""
        # We create summary even if some products failed, but log warnings
        failed_products = [r for r in results if r.status == "error" or not r.vendor_offers]
        if failed_products:
            logger.warning(f"{len(failed_products)} products had no results or errors")
        
        # Only skip summary if ALL products failed
        successful_products = [r for r in results if r.vendor_offers]
        return len(successful_products) > 0
    
    def _create_empty_summary_sheet(self, workbook: Workbook) -> None:
        """Create empty summary sheet when no products were scraped successfully."""
        ws = workbook.create_sheet("סיכום")
        
        # Hebrew headers
        headers = [
            "מק\"ט", "שם המוצר", "מחיר מקורי", "ספק הזול ביותר", 
            "מחיר הזול ביותר", "חיסכון", "% הפרש מ-ZAP", "קישור לספק"
        ]
        
        # Add headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        
        # Add a note about no results
        ws.cell(row=2, column=1, value="אין תוצאות - כל המוצרים נכשלו בגרידה")
        
        # Adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
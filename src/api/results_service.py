"""
Results Service - API interface for viewing and managing scraping results

This service provides a clean API interface to results management,
enabling CLI to access Excel files and processing history.
"""

import os
import glob
from typing import List, Dict, Any, Optional, Tuple
from pathlib import Path
from datetime import datetime
import openpyxl


class ResultsService:
    """Service class for results management operations."""
    
    def __init__(self):
        """Initialize results service."""
        self.project_root = Path(__file__).parent.parent.parent
        self.output_dir = self.project_root / "output"
        
        # Ensure output directory exists
        self.output_dir.mkdir(exist_ok=True)
    
    def get_recent_results(self, limit: int = 10) -> List[Dict[str, Any]]:
        """
        Get recent scraping results.
        
        Args:
            limit: Maximum number of results to return
            
        Returns:
            List of result information dictionaries
        """
        try:
            pattern = str(self.output_dir / "Lines_*_Report_*.xlsx")
            files = glob.glob(pattern)
            
            # Sort by modification time, newest first
            files.sort(key=lambda x: os.path.getmtime(x), reverse=True)
            
            results = []
            for file_path in files[:limit]:
                try:
                    result_info = self._analyze_excel_file(file_path)
                    results.append(result_info)
                except Exception as e:
                    # If we can't analyze a file, still include basic info
                    results.append({
                        "file_path": file_path,
                        "filename": os.path.basename(file_path),
                        "created_time": datetime.fromtimestamp(os.path.getctime(file_path)),
                        "file_size": os.path.getsize(file_path),
                        "analysis_error": str(e)
                    })
            
            return results
            
        except Exception as e:
            raise Exception(f"Failed to get recent results: {e}")
    
    def _analyze_excel_file(self, file_path: str) -> Dict[str, Any]:
        """Analyze Excel file and extract key information."""
        try:
            workbook = openpyxl.load_workbook(file_path, read_only=True)
            
            analysis = {
                "file_path": file_path,
                "filename": os.path.basename(file_path),
                "created_time": datetime.fromtimestamp(os.path.getctime(file_path)),
                "modified_time": datetime.fromtimestamp(os.path.getmtime(file_path)),
                "file_size": os.path.getsize(file_path),
                "worksheets": list(workbook.sheetnames),
                "rows_processed": self._extract_rows_from_filename(os.path.basename(file_path))
            }
            
            # Analyze פירוט (Details) worksheet if available
            if "פירוט" in workbook.sheetnames:
                details_sheet = workbook["פירוט"]
                analysis["total_vendors"] = details_sheet.max_row - 1  # Subtract header row
                
                # Try to extract price information
                price_data = self._extract_price_data(details_sheet)
                analysis.update(price_data)
            
            # Analyze סיכום (Summary) worksheet if available
            if "סיכום" in workbook.sheetnames:
                summary_sheet = workbook["סיכום"]
                summary_data = self._extract_summary_data(summary_sheet)
                analysis.update(summary_data)
            
            workbook.close()
            return analysis
            
        except Exception as e:
            raise Exception(f"Failed to analyze Excel file {file_path}: {e}")
    
    def _extract_price_data(self, sheet) -> Dict[str, Any]:
        """Extract price information from details sheet."""
        try:
            prices = []
            
            # Assuming price columns are around G-I (ZAP price, difference, etc.)
            for row in range(2, min(sheet.max_row + 1, 100)):  # Limit to avoid huge files
                try:
                    # Try to find price cells (looking for ₪ symbol or numeric values)
                    for col in range(6, 12):  # Columns F-K
                        cell_value = sheet.cell(row=row, column=col).value
                        if cell_value and isinstance(cell_value, (int, float)) and cell_value > 0:
                            prices.append(float(cell_value))
                            break
                except:
                    continue
            
            if prices:
                return {
                    "min_price": min(prices),
                    "max_price": max(prices),
                    "avg_price": sum(prices) / len(prices),
                    "price_samples": len(prices)
                }
            else:
                return {"price_analysis": "No price data found"}
                
        except Exception:
            return {"price_analysis": "Price extraction failed"}
    
    def _extract_summary_data(self, sheet) -> Dict[str, Any]:
        """Extract information from summary sheet."""
        try:
            summary_data = {}
            
            # Read first few rows for summary statistics
            for row in range(1, min(sheet.max_row + 1, 5)):
                for col in range(1, min(sheet.max_column + 1, 10)):
                    cell_value = sheet.cell(row=row, column=col).value
                    if cell_value and isinstance(cell_value, str):
                        # Look for key metrics
                        if "vendor" in cell_value.lower():
                            try:
                                next_cell = sheet.cell(row=row, column=col+1).value
                                if isinstance(next_cell, (int, float)):
                                    summary_data["summary_vendor_count"] = int(next_cell)
                            except:
                                pass
                        elif "success" in cell_value.lower() or "rate" in cell_value.lower():
                            try:
                                next_cell = sheet.cell(row=row, column=col+1).value
                                if isinstance(next_cell, (int, float)):
                                    summary_data["success_rate"] = float(next_cell)
                            except:
                                pass
            
            return summary_data
            
        except Exception:
            return {"summary_analysis": "Summary extraction failed"}
    
    def _extract_rows_from_filename(self, filename: str) -> str:
        """Extract row information from Excel filename."""
        import re
        # Pattern: Lines_126_Report_... or Lines_126-127_Report_...
        match = re.search(r'Lines_([0-9-]+)_Report_', filename)
        if match:
            return match.group(1)
        return "unknown"
    
    def open_excel_file(self, file_path: str) -> bool:
        """
        Open Excel file in default application.
        
        Args:
            file_path: Path to Excel file
            
        Returns:
            True if successfully opened
        """
        try:
            if not os.path.exists(file_path):
                return False
            
            import subprocess
            import sys
            
            if sys.platform == "win32":
                os.startfile(file_path)
            elif sys.platform == "darwin":
                subprocess.run(["open", file_path])
            else:
                subprocess.run(["xdg-open", file_path])
            
            return True
            
        except Exception:
            return False
    
    def get_file_details(self, file_path: str) -> Dict[str, Any]:
        """
        Get detailed information about a specific Excel file.
        
        Args:
            file_path: Path to Excel file
            
        Returns:
            Detailed file information
        """
        try:
            if not os.path.exists(file_path):
                return {"error": "File not found"}
            
            return self._analyze_excel_file(file_path)
            
        except Exception as e:
            return {"error": f"Failed to get file details: {e}"}
    
    def search_results(self, search_term: str) -> List[Dict[str, Any]]:
        """
        Search for results containing specific term.
        
        Args:
            search_term: Term to search for in filenames or content
            
        Returns:
            List of matching results
        """
        try:
            all_results = self.get_recent_results(limit=50)
            
            matching_results = []
            search_lower = search_term.lower()
            
            for result in all_results:
                # Search in filename
                if search_lower in result["filename"].lower():
                    matching_results.append(result)
                    continue
                
                # Search in rows processed
                if "rows_processed" in result and search_lower in str(result["rows_processed"]).lower():
                    matching_results.append(result)
                    continue
            
            return matching_results
            
        except Exception as e:
            raise Exception(f"Failed to search results: {e}")
    
    def get_statistics(self) -> Dict[str, Any]:
        """
        Get overall statistics about scraping results.
        
        Returns:
            Statistics dictionary
        """
        try:
            all_results = self.get_recent_results(limit=100)
            
            if not all_results:
                return {"message": "No results found"}
            
            stats = {
                "total_files": len(all_results),
                "date_range": {
                    "earliest": min(r["created_time"] for r in all_results if "created_time" in r),
                    "latest": max(r["created_time"] for r in all_results if "created_time" in r)
                },
                "total_vendors_processed": 0,
                "average_vendors_per_file": 0,
                "total_file_size": sum(r.get("file_size", 0) for r in all_results)
            }
            
            # Calculate vendor statistics
            vendor_counts = [r["total_vendors"] for r in all_results if "total_vendors" in r]
            if vendor_counts:
                stats["total_vendors_processed"] = sum(vendor_counts)
                stats["average_vendors_per_file"] = sum(vendor_counts) / len(vendor_counts)
                stats["max_vendors_in_file"] = max(vendor_counts)
                stats["min_vendors_in_file"] = min(vendor_counts)
            
            # Price statistics
            prices = []
            for result in all_results:
                if "min_price" in result and "max_price" in result:
                    prices.extend([result["min_price"], result["max_price"]])
            
            if prices:
                stats["price_range"] = {
                    "lowest": min(prices),
                    "highest": max(prices),
                    "average": sum(prices) / len(prices)
                }
            
            return stats
            
        except Exception as e:
            raise Exception(f"Failed to get statistics: {e}")